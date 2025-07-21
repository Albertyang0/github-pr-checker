# === MODULE IMPORTS WITH DEPENDENCY CHECK ===
try:
    import asyncio
    import aiohttp
    import openpyxl
    import os
    import time
    import subprocess
    import re
    from datetime import datetime
    from openpyxl.styles import Font, PatternFill
    from rich.progress import Progress, SpinnerColumn, BarColumn, TimeElapsedColumn
except ImportError as e:
    print(f"Missing module: {e.name}. Please install it using pip.")
    exit(1)

# === USER CONFIGURATION ===
GITHUB_TOKEN = "GITHUB_TOKEN"  # Set your GitHub personal access token
INPUT_PATH = r"C:\\Users\\v-bowenyang\\Desktop\\PR_check.xlsx"  # Excel file with PR URLs

DATE_FOLDER = datetime.now().strftime("%Y-%m-%d")
OUTPUT_BASE_DIR = os.path.join(os.path.expanduser("~"), "Desktop", "PR_Check_Results", DATE_FOLDER)
API_DELAY = 0.1  # Delay between API calls to avoid rate limits
CONCURRENT_REQUESTS = 10  # Maximum concurrent GitHub API requests
KEEP_LAST_N_RUNS = 10  # Number of past run folders to retain

# === GLOBAL VARIABLES ===
start_time = time.time()
timestamp_str = datetime.now().strftime("%H%M%S")
output_dir = None
log_file = None

# Filter settings
excluded_users = {"acrolinx-at-msft2", "learn-build-service-prod"}
excluded_comment_keywords = ["Acrolinx Scorecards"]

MERGEABLE_STATE_MAPPING = {
    "dirty": "Has conflicts",
    "blocked": "Blocked",
    "behind": "Behind base branch",
    "unstable": "Checks failed",
    "unknown": "Unknown state",
    "clean": "Mergeable"
}

# === HELPER FUNCTIONS ===

# Determine if a GitHub user is a bot
def is_bot(user):
    login = user.get("login", "").lower()
    return user.get("type") == "Bot" or login.endswith("[bot]") or "bot" in login

# Clean illegal characters for Excel compatibility
def clean_illegal_chars(text):
    if isinstance(text, str):
        return re.sub(r"[\x00-\x1F\x7F-\x9F]", "", text)
    return text

# Write log entry to file
def write_log_entry(pr_url, status, external_comment):
    if log_file:
        log_file.write(f"[{datetime.now()}] {pr_url} → {status}, External: {external_comment}\n")

# Create a new output folder and remove older runs
def create_output_dir():
    global output_dir, log_file
    os.makedirs(OUTPUT_BASE_DIR, exist_ok=True)
    output_dir = os.path.join(OUTPUT_BASE_DIR, f"Run_{timestamp_str}")
    os.makedirs(output_dir, exist_ok=True)

    # Clean up older folders
    run_folders = sorted([
        f for f in os.listdir(OUTPUT_BASE_DIR)
        if os.path.isdir(os.path.join(OUTPUT_BASE_DIR, f)) and f.startswith("Run_")
    ])
    while len(run_folders) > KEEP_LAST_N_RUNS:
        folder_to_delete = os.path.join(OUTPUT_BASE_DIR, run_folders.pop(0))
        try:
            subprocess.call(["rmdir", "/s", "/q", folder_to_delete], shell=True)
        except Exception as e:
            print(f"Warning: Failed to delete old folder: {folder_to_delete}: {e}")

    log_path = os.path.join(output_dir, "PR_check_log.txt")
    log_file = open(log_path, "w", encoding="utf-8")

# Auto-fit column width in Excel worksheet
def auto_adjust_column_width(ws):
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[column].width = max_length + 2

# Check if PR is older than threshold, unmerged, and has no conflicts.
def is_old_unmerged_no_conflicts(created_at, merged_status, mergeable_state, days_threshold=7):
    try:
        if created_at:
            from datetime import datetime, timezone
            pr_create_time = datetime.strptime(created_at, "%Y-%m-%dT%H:%M:%SZ").replace(tzinfo=timezone.utc)
            days_old = (datetime.now(timezone.utc) - pr_create_time).days
            if (
                days_old > days_threshold and
                not merged_status.startswith("Merged") and
                mergeable_state == "clean"
            ):
                return "Yes"
            else:
                return "No"
        else:
            return "No"
    except Exception:
        return "Error"

# Print a simple ASCII table with auto-adjusted column widths
def print_table(headers, rows):
    all_rows = [headers] + rows
    col_widths = [max(len(str(row[i])) for row in all_rows) for i in range(len(headers))]

    def print_line():
        print("+" + "+".join("-" * (w + 2) for w in col_widths) + "+")

    def print_row(row):
        formatted_cells = []
        for i, val in enumerate(row):
            cell = str(val)
            if i == 0:
                formatted_cells.append(cell.ljust(col_widths[i]))  # Left-align first column
            else:
                formatted_cells.append(cell.rjust(col_widths[i]))  # Right-align other columns
        print("| " + " | ".join(formatted_cells) + " |")

    print_line()
    print_row(headers)
    print_line()
    for row in rows:
        print_row(row)
    print_line()

# === ASYNC TASK: Fetch PR status and comments ===

async def fetch_status(session, sem, pr_url, retries=2, retry_delay=2):
    """
    Fetch PR metadata and comments from GitHub with automatic retries
    if mergeable_state is 'unknown'.

    Args:
        session: aiohttp ClientSession instance
        sem: asyncio.Semaphore for concurrency control
        pr_url: PR URL string
        retries: Number of retries if mergeable_state is 'unknown' (default 2)
        retry_delay: Seconds to wait between retries (default 2)

    Returns:
        Tuple with PR URL, merged status, external comment flag, external comments, author login
    """
    async with sem:
        try:
            # Parse PR URL
            parts = pr_url.split("/")
            if len(parts) < 7:
                return pr_url, "Invalid URL", "Skipped", "Skipped", "Unknown"

            owner, repo, pr_number = parts[3], parts[4], parts[6]
            pr_api = f"https://api.github.com/repos/{owner}/{repo}/pulls/{pr_number}"
            comment_api = f"https://api.github.com/repos/{owner}/{repo}/issues/{pr_number}/comments?per_page=100"

            # === Retry logic for mergeable_state ===
            mergeable_state = None
            pr_data = None
            for attempt in range(retries):
                async with session.get(pr_api) as pr_resp:
                    if pr_resp.status != 200:
                        raise Exception(f"Failed to fetch PR info. Status code: {pr_resp.status}")
                    pr_data = await pr_resp.json()
                    mergeable_state = pr_data.get("mergeable_state")
                    # If the state is not 'unknown', break and use this value
                    if mergeable_state != "unknown":
                        break
                    await asyncio.sleep(retry_delay)  # Wait before retrying

            # Determine PR status based on API response
            if pr_data.get("state") == "closed" and not pr_data.get("merged"):
                status = "Closed (Not merged)"
            else:
                status = "Merged" if pr_data.get("merged") else "Not merged"
                if mergeable_state and not pr_data.get("merged"):
                    status += f" ({MERGEABLE_STATE_MAPPING.get(mergeable_state, mergeable_state)})"
            author_login = pr_data.get("user", {}).get("login", "Unknown")

            await asyncio.sleep(API_DELAY)  # Respect API rate limits

            # === Fetch PR comments ===
            async with session.get(comment_api) as cmt_resp:
                if cmt_resp.status != 200:
                    raise Exception(f"Failed to fetch comments. Status code: {cmt_resp.status}")
                comments = await cmt_resp.json()

            # Filter external comments (not from excluded users, not bots, no excluded keywords)
            external_comments = [
                c.get("body", "") for c in comments
                if (login := c.get("user", {}).get("login"))
                and login not in excluded_users
                and not is_bot(c.get("user", {}))
                and not any(keyword in c.get("body", "") for keyword in excluded_comment_keywords)
            ]

            external_flag = "Yes" if external_comments else "No"
            comment_text = "\n---\n".join(external_comments) if external_comments else "None"

            write_log_entry(pr_url, status, external_flag)
            created_at = pr_data.get("created_at")
            return pr_url, status, external_flag, comment_text, author_login, created_at, mergeable_state
        except Exception as e:
            print(f"⚠️ Error occurred while processing {pr_url}: {e}")
            return pr_url, "Error", "Error", str(e), "Unknown", None, None

# === ASYNC MAIN: Process all PRs in Excel ===

async def process_all_prs(session):
    wb = openpyxl.load_workbook(INPUT_PATH)
    ws = wb.active
    font = Font(name="DengXian", size=11)
    green_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    red_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    gray_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")

    # Determine new columns for output
    col_author = ws.max_column + 1
    col_status = col_author + 1
    col_old_unmerged_no_conflicts = col_status + 1
    col_external_content = col_old_unmerged_no_conflicts + 1

    # Set headers
    ws.cell(row=1, column=1, value="PR URL").font = font
    ws.cell(row=1, column=col_author, value="Author").font = font
    ws.cell(row=1, column=col_status, value="Merged Status").font = font
    ws.cell(row=1, column=col_old_unmerged_no_conflicts, value="Old & Unmerged & No Conflicts").font = font
    ws.cell(row=1, column=col_external_content, value="External Comments Content").font = font

    sem = asyncio.Semaphore(CONCURRENT_REQUESTS)
    rows = list(ws.iter_rows(min_row=2))

    unique_prs = {}
    row_to_pr = {}

    for i, row in enumerate(rows, start=2):
        pr_url = row[0].value
        if pr_url and "github.com" in pr_url:
            unique_prs[pr_url] = None
            row_to_pr[i] = pr_url
        else:
            for col, value in zip(
                [col_author, col_status, col_old_unmerged_no_conflicts, col_external_content],
                ["N/A"] * 4):
                ws.cell(row=i, column=col, value=value).font = font
                ws.cell(row=i, column=col).fill = gray_fill

    if output_dir is None:
        create_output_dir()

    # Launch async fetch tasks
    pr_tasks = [fetch_status(session, sem, pr_url) for pr_url in unique_prs.keys()]

    with Progress(
        SpinnerColumn(),
        "[progress.description]{task.description}",
        BarColumn(),
        "[progress.percentage]{task.percentage:>3.0f}%",
        TimeElapsedColumn(),
    ) as progress:
        task_id = progress.add_task("Processing PRs...", total=len(pr_tasks))
        results = []
        for coro in asyncio.as_completed(pr_tasks):
            result = await coro
            results.append(result)
            progress.update(task_id, advance=1)

    # Store results
    for pr_url, status, external_flag, content, author_login, created_at, mergeable_state in results:
        unique_prs[pr_url] = {
            "merged_status": status,
            "has_external_comment": external_flag,
            "external_comments": content,
            "author": author_login,
            "created_at": created_at,
            "mergeable_state": mergeable_state
        }

    # Write results to Excel
    for row_idx, pr_url in row_to_pr.items():
        result = unique_prs.get(pr_url)
        old_unmerged_no_conflicts_flag = is_old_unmerged_no_conflicts(
            result.get("created_at"),
            result.get("merged_status"),
            result.get("mergeable_state")
        )
        if result:
            ws.cell(row=row_idx, column=col_author, value=result["author"]).font = font
            ws.cell(row=row_idx, column=col_status, value=result["merged_status"]).font = font
            ws.cell(row=row_idx, column=col_status).fill = green_fill if result["merged_status"].startswith("Merged") else red_fill
            ws.cell(row=row_idx, column=col_old_unmerged_no_conflicts, value=old_unmerged_no_conflicts_flag).font = font
            ws.cell(row=row_idx, column=col_external_content, value=clean_illegal_chars(result["external_comments"])).font = font

            # Save the flag to unique_prs
            result["old_unmerged_no_conflicts_flag"] = old_unmerged_no_conflicts_flag

    # Save file and output summary
    if output_dir:
        result_path = os.path.join(output_dir, "PR_Check_Result_Async.xlsx")
        auto_adjust_column_width(ws)
        wb.save(result_path)

        print(f"\n✅ Done! Results saved to: {output_dir}")
        os.startfile(result_path)

        # === STATISTICS OUTPUT ===
        old_unmerged_no_conflicts_count = sum(1 for v in unique_prs.values() if v and v.get("old_unmerged_no_conflicts_flag") == "Yes")
        merged_count = sum(1 for v in unique_prs.values() if v and v["merged_status"].startswith("Merged"))
        total_count = len(unique_prs)

        # Assemble the summary data
        summary = [
            ["Total PRs", total_count],
            ["Merged PRs", merged_count],
            ["PRs open for more than 1 week", old_unmerged_no_conflicts_count]
        ]
        headers = ["Metric", "Value"]

        print("\n📊 Summary Statistics:\n")
        print_table(headers, summary)
    else:
        print("\n⚠️ No PRs processed, no output generated.")

# === MAIN ENTRY POINT ===
if __name__ == "__main__":
    async def run():
        headers = {
            "Authorization": f"token {GITHUB_TOKEN}",
            "Accept": "application/vnd.github+json"
        }
        async with aiohttp.ClientSession(headers=headers) as session:
            await process_all_prs(session)

    try:
        loop = asyncio.new_event_loop()
        asyncio.set_event_loop(loop)
        loop.run_until_complete(run())
    except Exception as e:
        print(f"❌ Unexpected error: {e}")
    finally:
        loop.close()
        end_time = time.time()
        print(f"⏱️ Execution time: {end_time - start_time:.2f} seconds")
        print("\n")
